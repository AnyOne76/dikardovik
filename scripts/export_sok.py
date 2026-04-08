from __future__ import annotations

import glob
import html
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    desktop = Path.home() / "Desktop"
    candidates = sorted(glob.glob(str(desktop / "*.xlsx")))
    if not candidates:
        raise SystemExit(f"No .xlsx files found in {desktop}")

    src = Path(candidates[0])
    out = Path(__file__).resolve().parents[1] / "public" / "sok-sample-export.html"

    wb = load_workbook(src, data_only=True)
    parts: list[str] = [
        "<!doctype html>",
        "<meta charset='utf-8'>",
        "<title>SOK export</title>",
        "<style>body{font-family:Arial,Helvetica,sans-serif} table{border-collapse:collapse} td{vertical-align:top;max-width:520px;white-space:normal}</style>",
    ]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"<h2>{html.escape(sheet_name)}</h2>")
        parts.append("<table border='1' cellspacing='0' cellpadding='4'>")
        for r in range(1, ws.max_row + 1):
            parts.append("<tr>")
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                s = "" if v is None else str(v)
                s = html.escape(s).replace("\n", "<br>")
                parts.append(f"<td>{s}</td>")
            parts.append("</tr>")
        parts.append("</table>")

    out.write_text("\n".join(parts), encoding="utf-8")
    print(str(out))


if __name__ == "__main__":
    main()

